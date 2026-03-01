#!/usr/bin/env python3
"""Compatibility-first Excel interoperability verification harness.

This script validates bidirectional XLSX interoperability:
1) Excel/openpyxl-authored workbook -> RootCellar CLI open/save/recalc paths.
2) RootCellar-produced workbooks -> openpyxl load/save paths.
3) Optional corpus sweep over multiple .xlsx files with deterministic ordering.
"""

from __future__ import annotations

import argparse
import json
import re
import shutil
import subprocess
import sys
import zipfile
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook


@dataclass
class CommandRecord:
    case: str
    step: str
    command: list[str]
    exit_code: int
    stdout: str
    stderr: str


@dataclass
class CaseReport:
    name: str
    source_workbook: str
    mutation_sheet: str
    checks: list[str]
    artifacts: dict[str, str]


def run_command(command: list[str], cwd: Path, case: str, step: str) -> CommandRecord:
    completed = subprocess.run(
        command,
        cwd=str(cwd),
        capture_output=True,
        text=True,
        check=False,
    )
    record = CommandRecord(
        case=case,
        step=step,
        command=command,
        exit_code=completed.returncode,
        stdout=completed.stdout,
        stderr=completed.stderr,
    )
    if record.exit_code != 0:
        rendered = " ".join(command)
        raise RuntimeError(
            f"command failed ({record.exit_code}): {rendered}\n"
            f"stdout:\n{record.stdout}\n"
            f"stderr:\n{record.stderr}"
        )
    return record


def assert_zip_integrity(path: Path) -> None:
    with zipfile.ZipFile(path, "r") as archive:
        bad_entry = archive.testzip()
        if bad_entry is not None:
            raise RuntimeError(f"zip integrity check failed for {path}: bad entry {bad_entry}")


def assert_workbook_loadable(path: Path) -> None:
    workbook = load_workbook(path, data_only=False)
    workbook.close()


def write_excel_fixture(path: Path, sheet_name: str) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_name
    sheet["A1"] = 10
    sheet["B1"] = 20
    sheet["C1"] = "=A1+B1"
    sheet["D1"] = "hello"
    sheet["E1"] = True
    workbook.save(path)


def assert_semantics(path: Path, sheet_name: str, expected_a1: Any) -> None:
    workbook = load_workbook(path, data_only=False)
    try:
        if sheet_name not in workbook.sheetnames:
            raise RuntimeError(f"missing sheet {sheet_name!r} in {path}")
        sheet = workbook[sheet_name]
        if sheet["A1"].value != expected_a1:
            raise RuntimeError(
                f"unexpected {sheet_name}!A1 in {path}: "
                f"expected {expected_a1!r}, got {sheet['A1'].value!r}"
            )
        if sheet["C1"].value != "=A1+B1":
            raise RuntimeError(
                f"unexpected {sheet_name}!C1 formula in {path}: "
                f"expected '=A1+B1', got {sheet['C1'].value!r}"
            )
        if sheet["D1"].value != "hello":
            raise RuntimeError(
                f"unexpected {sheet_name}!D1 text in {path}: "
                f"expected 'hello', got {sheet['D1'].value!r}"
            )
        if sheet["E1"].value is not True:
            raise RuntimeError(
                f"unexpected {sheet_name}!E1 bool in {path}: "
                f"expected True, got {sheet['E1'].value!r}"
            )
    finally:
        workbook.close()


def assert_mutated_a1(path: Path, sheet_name: str, expected_a1: Any) -> None:
    workbook = load_workbook(path, data_only=False)
    try:
        if sheet_name not in workbook.sheetnames:
            raise RuntimeError(f"missing sheet {sheet_name!r} in {path}")
        sheet = workbook[sheet_name]
        if sheet["A1"].value != expected_a1:
            raise RuntimeError(
                f"unexpected {sheet_name}!A1 in {path}: "
                f"expected {expected_a1!r}, got {sheet['A1'].value!r}"
            )
    finally:
        workbook.close()


def sanitize_case_name(name: str) -> str:
    compact = re.sub(r"[^a-zA-Z0-9._-]+", "_", name).strip("_")
    return compact or "case"


def discover_xlsx_files(corpus_dir: Path) -> list[Path]:
    if not corpus_dir.is_dir():
        raise RuntimeError(f"corpus directory does not exist: {corpus_dir}")
    files = sorted(
        [p for p in corpus_dir.rglob("*") if p.is_file() and p.suffix.lower() == ".xlsx"],
        key=lambda p: p.as_posix().lower(),
    )
    if not files:
        raise RuntimeError(f"no .xlsx files found in corpus directory: {corpus_dir}")
    return files


def load_manifest(manifest_path: Path) -> dict[str, Any]:
    try:
        payload = json.loads(manifest_path.read_text(encoding="utf-8"))
    except Exception as exc:
        raise RuntimeError(f"failed to read corpus manifest {manifest_path}: {exc}") from exc
    if not isinstance(payload, dict):
        raise RuntimeError(f"invalid corpus manifest format at {manifest_path}: expected object")
    return payload


def select_mutation_sheet(path: Path, preferred_sheet: str) -> str:
    workbook = load_workbook(path, data_only=False)
    try:
        if preferred_sheet in workbook.sheetnames:
            return preferred_sheet
        if not workbook.sheetnames:
            raise RuntimeError(f"workbook has no sheets: {path}")
        return workbook.sheetnames[0]
    finally:
        workbook.close()


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Verify bidirectional Excel interoperability for RootCellar XLSX flows."
    )
    parser.add_argument(
        "--workspace",
        default=".",
        help="Workspace root containing Cargo.toml (default: current directory).",
    )
    parser.add_argument(
        "--workdir",
        default="./target/excel-interop-gate",
        help="Scratch/output directory for generated artifacts.",
    )
    parser.add_argument(
        "--input",
        default=None,
        help="Optional existing XLSX workbook. If omitted, a fixture workbook is generated.",
    )
    parser.add_argument(
        "--corpus-dir",
        default=None,
        help="Optional directory containing .xlsx files to run through the same interop checks.",
    )
    parser.add_argument(
        "--corpus-manifest",
        default=None,
        help=(
            "Optional fixture manifest path. If omitted and "
            "<corpus-dir>/manifest.json exists, it is included automatically."
        ),
    )
    parser.add_argument(
        "--max-corpus-files",
        type=int,
        default=0,
        help="Optional max corpus files to process after deterministic ordering (0 = all).",
    )
    parser.add_argument(
        "--require-corpus-fixture",
        action="append",
        default=[],
        help=(
            "Require a fixture basename to exist in selected corpus files. "
            "May be supplied multiple times."
        ),
    )
    parser.add_argument(
        "--sheet",
        default="Sheet1",
        help="Sheet name used for tx-save mutation checks (default: Sheet1).",
    )
    parser.add_argument(
        "--report",
        default="./target/excel-interop-gate-report.json",
        help="Path to write JSON verification report.",
    )
    parser.add_argument(
        "--cargo-offline",
        action="store_true",
        help="Pass --offline to cargo run commands (default: disabled).",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    workspace = Path(args.workspace).resolve()
    workdir = Path(args.workdir).resolve()
    report_path = Path(args.report).resolve()
    workdir.mkdir(parents=True, exist_ok=True)

    command_records: list[CommandRecord] = []
    checks: list[str] = []
    case_reports: list[CaseReport] = []
    cases_dir = workdir / "cases"
    cases_dir.mkdir(parents=True, exist_ok=True)

    def run_cli(case_name: str, step: str, cli_args: list[str]) -> None:
        command = ["cargo", "run", "-p", "rootcellar-cli", "--locked"]
        if args.cargo_offline:
            command.append("--offline")
        command.extend(["--", *cli_args])
        record = run_command(command, workspace, case_name, step)
        command_records.append(record)

    generated_fixture = args.input is None
    primary_case_name = "primary"
    primary_case_dir = cases_dir / primary_case_name
    primary_case_dir.mkdir(parents=True, exist_ok=True)
    primary_source = primary_case_dir / "source.xlsx"
    if generated_fixture:
        write_excel_fixture(primary_source, args.sheet)
    else:
        input_path = Path(args.input).resolve()
        if not input_path.is_file():
            raise RuntimeError(f"input workbook does not exist: {input_path}")
        shutil.copy2(input_path, primary_source)

    source_cases: list[tuple[str, Path, bool]] = [(primary_case_name, primary_source, generated_fixture)]

    corpus_sources: list[str] = []
    corpus_manifest_path: Path | None = None
    corpus_manifest: dict[str, Any] | None = None
    missing_required_fixtures: list[str] = []
    if args.corpus_dir:
        corpus_root = Path(args.corpus_dir).resolve()
        discovered = discover_xlsx_files(corpus_root)
        if args.max_corpus_files > 0:
            discovered = discovered[: args.max_corpus_files]

        selected_names = {source.name for source in discovered}
        missing_required_fixtures = sorted(
            {name for name in args.require_corpus_fixture if name not in selected_names}
        )
        if missing_required_fixtures:
            raise RuntimeError(
                "required corpus fixtures missing from selected corpus set: "
                f"{missing_required_fixtures}; selected={sorted(selected_names)}"
            )

        if args.corpus_manifest:
            corpus_manifest_path = Path(args.corpus_manifest).resolve()
            if not corpus_manifest_path.is_file():
                raise RuntimeError(
                    f"corpus manifest does not exist: {corpus_manifest_path}"
                )
        else:
            default_manifest_path = corpus_root / "manifest.json"
            if default_manifest_path.is_file():
                corpus_manifest_path = default_manifest_path

        if corpus_manifest_path is not None:
            corpus_manifest = load_manifest(corpus_manifest_path)

        for idx, source in enumerate(discovered, start=1):
            rel = source.relative_to(corpus_root).as_posix()
            corpus_sources.append(rel)
            case_name = f"corpus-{idx:03d}-{sanitize_case_name(rel)}"
            case_dir = cases_dir / case_name
            case_dir.mkdir(parents=True, exist_ok=True)
            case_source = case_dir / "source.xlsx"
            shutil.copy2(source, case_source)
            source_cases.append((case_name, case_source, False))

    def run_case(case_name: str, source: Path, expect_fixture_semantics: bool) -> CaseReport:
        case_dir = source.parent
        preserved_path = case_dir / "rootcellar-preserved.xlsx"
        normalized_path = case_dir / "rootcellar-normalized.xlsx"
        mutated_path = case_dir / "rootcellar-mutated.xlsx"
        excel_roundtrip_path = case_dir / "excel-roundtrip.xlsx"

        case_checks: list[str] = []
        mutation_sheet = select_mutation_sheet(source, args.sheet)

        run_cli(case_name, "open_source", ["open", str(source)])
        run_cli(
            case_name,
            "save_preserve",
            ["save", str(source), str(preserved_path), "--mode", "preserve"],
        )
        run_cli(
            case_name,
            "save_normalize",
            ["save", str(source), str(normalized_path), "--mode", "normalize"],
        )
        run_cli(
            case_name,
            "tx_save_mutate_a1",
            [
                "tx-save",
                str(source),
                str(mutated_path),
                "--sheet",
                mutation_sheet,
                "--set",
                "A1=99",
                "--mode",
                "preserve",
            ],
        )
        case_checks.append("rootcellar_open_save_txsave_passed")

        for path in [source, preserved_path, normalized_path, mutated_path]:
            assert_zip_integrity(path)
            assert_workbook_loadable(path)
        case_checks.append("zip_integrity_and_openpyxl_load_passed")

        if expect_fixture_semantics:
            assert_semantics(source, args.sheet, 10)
            assert_semantics(preserved_path, args.sheet, 10)
            assert_semantics(normalized_path, args.sheet, 10)
            case_checks.append("fixture_semantics_passed")

        assert_mutated_a1(mutated_path, mutation_sheet, 99)
        case_checks.append("tx_save_mutation_verified")

        workbook = load_workbook(preserved_path, data_only=False)
        workbook.save(excel_roundtrip_path)
        workbook.close()
        assert_zip_integrity(excel_roundtrip_path)
        assert_workbook_loadable(excel_roundtrip_path)
        case_checks.append("excel_roundtrip_save_passed")

        run_cli(case_name, "open_roundtrip", ["open", str(excel_roundtrip_path)])
        run_cli(
            case_name,
            "recalc_roundtrip",
            ["recalc", str(excel_roundtrip_path), "--sheet", mutation_sheet],
        )
        case_checks.append("rootcellar_open_recalc_roundtrip_passed")

        return CaseReport(
            name=case_name,
            source_workbook=str(source),
            mutation_sheet=mutation_sheet,
            checks=case_checks,
            artifacts={
                "preserved_workbook": str(preserved_path),
                "normalized_workbook": str(normalized_path),
                "mutated_workbook": str(mutated_path),
                "excel_roundtrip_workbook": str(excel_roundtrip_path),
            },
        )

    for case_name, source, expect_fixture_semantics in source_cases:
        report = run_case(case_name, source, expect_fixture_semantics)
        case_reports.append(report)
        checks.append(f"{case_name}:pass")

    report = {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "workspace": str(workspace),
        "workdir": str(workdir),
        "generated_fixture": generated_fixture,
        "corpus_sources": corpus_sources,
        "corpus_required_fixtures": sorted(set(args.require_corpus_fixture)),
        "corpus_missing_required_fixtures": missing_required_fixtures,
        "corpus_manifest_path": str(corpus_manifest_path)
        if corpus_manifest_path is not None
        else None,
        "corpus_manifest": corpus_manifest,
        "checks": checks,
        "cases": [
            {
                "name": case.name,
                "source_workbook": case.source_workbook,
                "mutation_sheet": case.mutation_sheet,
                "checks": case.checks,
                "artifacts": case.artifacts,
            }
            for case in case_reports
        ],
        "commands": [
            {
                "case": record.case,
                "step": record.step,
                "command": record.command,
                "exit_code": record.exit_code,
                "stdout": record.stdout,
                "stderr": record.stderr,
            }
            for record in command_records
        ],
        "artifact_case_dir": str(cases_dir),
        "status": "pass",
    }
    report_path.parent.mkdir(parents=True, exist_ok=True)
    report_path.write_text(json.dumps(report, indent=2), encoding="utf-8")
    print(f"Excel interop verification report: {report_path}")
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except Exception as exc:  # pragma: no cover - script-level guard
        print(f"excel interop verification failed: {exc}", file=sys.stderr)
        raise SystemExit(1)
