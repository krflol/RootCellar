#!/usr/bin/env python3
"""Generate deterministic XLSX fixtures for corpus and interop validation."""

from __future__ import annotations

import argparse
import json
import pathlib
import zipfile
from dataclasses import dataclass
from datetime import datetime, timezone

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.workbook.defined_name import DefinedName


DETERMINISTIC_TIMESTAMP = datetime(2020, 1, 1, 0, 0, 0)


@dataclass(frozen=True)
class FixtureDescriptor:
    name: str
    path: pathlib.Path
    features: tuple[str, ...]
    generator: str


def build_content_types(include_shared_strings: bool) -> str:
    shared_override = (
        '<Override PartName="/xl/sharedStrings.xml" '
        'ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sharedStrings+xml"/>'
        if include_shared_strings
        else ""
    )
    return (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
        '<Default Extension="rels" '
        'ContentType="application/vnd.openxmlformats-package.relationships+xml"/>'
        '<Default Extension="xml" ContentType="application/xml"/>'
        '<Override PartName="/xl/workbook.xml" '
        'ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/>'
        '<Override PartName="/xl/worksheets/sheet1.xml" '
        'ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>'
        f"{shared_override}"
        "</Types>"
    )


def build_root_rels() -> str:
    return (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        '<Relationship Id="rId1" '
        'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" '
        'Target="xl/workbook.xml"/>'
        "</Relationships>"
    )


def build_workbook_xml(sheet_name: str = "Sheet1") -> str:
    return (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
        'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">'
        "<sheets>"
        f'<sheet name="{sheet_name}" sheetId="1" r:id="rId1"/>'
        "</sheets>"
        "</workbook>"
    )


def build_workbook_rels(extra_dangling: bool = False) -> str:
    dangling = (
        '<Relationship Id="rId2" '
        'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" '
        'Target="worksheets/missing-sheet.xml"/>'
        if extra_dangling
        else ""
    )
    return (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        '<Relationship Id="rId1" '
        'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" '
        'Target="worksheets/sheet1.xml"/>'
        f"{dangling}"
        "</Relationships>"
    )


def write_fixture(
    out_path: pathlib.Path,
    sheet_xml: str,
    *,
    include_shared_strings: bool = False,
    shared_strings_xml: str | None = None,
    extra_dangling: bool = False,
    extra_parts: dict[str, bytes] | None = None,
) -> None:
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with zipfile.ZipFile(out_path, mode="w", compression=zipfile.ZIP_DEFLATED) as archive:
        archive.writestr("[Content_Types].xml", build_content_types(include_shared_strings))
        archive.writestr("_rels/.rels", build_root_rels())
        archive.writestr("xl/workbook.xml", build_workbook_xml())
        archive.writestr("xl/_rels/workbook.xml.rels", build_workbook_rels(extra_dangling))
        archive.writestr("xl/worksheets/sheet1.xml", sheet_xml)
        if include_shared_strings and shared_strings_xml:
            archive.writestr("xl/sharedStrings.xml", shared_strings_xml)
        if extra_parts:
            for part_name, payload in extra_parts.items():
                archive.writestr(part_name, payload)


def apply_deterministic_metadata(workbook: Workbook, title: str) -> None:
    workbook.properties.creator = "RootCellar"
    workbook.properties.lastModifiedBy = "RootCellar"
    workbook.properties.title = title
    workbook.properties.created = DETERMINISTIC_TIMESTAMP
    workbook.properties.modified = DETERMINISTIC_TIMESTAMP


def write_styles_fixture(path: pathlib.Path) -> None:
    workbook = Workbook()
    apply_deterministic_metadata(workbook, "styles fixture")
    sheet = workbook.active
    sheet.title = "Sheet1"
    sheet["A1"] = 1234.5
    sheet["A2"] = "styled"
    sheet["A1"].number_format = "#,##0.00"
    sheet["A1"].font = Font(name="Calibri", bold=True, color="FF1F4E78")
    sheet["A1"].fill = PatternFill(fill_type="solid", fgColor="FFE2F0D9")
    sheet["A1"].alignment = Alignment(horizontal="center")
    thin = Side(style="thin", color="FF808080")
    sheet["A1"].border = Border(left=thin, right=thin, top=thin, bottom=thin)
    workbook.save(path)


def write_comments_fixture(path: pathlib.Path) -> None:
    workbook = Workbook()
    apply_deterministic_metadata(workbook, "comments fixture")
    sheet = workbook.active
    sheet.title = "Sheet1"
    sheet["A1"] = "needs review"
    sheet["A1"].comment = Comment("Compatibility comment fixture", "RootCellar")
    sheet["B1"] = 42
    workbook.save(path)


def write_chart_fixture(path: pathlib.Path) -> None:
    workbook = Workbook()
    apply_deterministic_metadata(workbook, "chart fixture")
    sheet = workbook.active
    sheet.title = "Sheet1"
    rows = [
        ("Quarter", "Revenue"),
        ("Q1", 12),
        ("Q2", 18),
        ("Q3", 9),
        ("Q4", 15),
    ]
    for row in rows:
        sheet.append(row)
    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = "Quarterly Revenue"
    chart.y_axis.title = "Revenue"
    chart.x_axis.title = "Quarter"
    values = Reference(sheet, min_col=2, min_row=1, max_row=len(rows))
    categories = Reference(sheet, min_col=1, min_row=2, max_row=len(rows))
    chart.add_data(values, titles_from_data=True)
    chart.set_categories(categories)
    sheet.add_chart(chart, "D2")
    workbook.save(path)


def write_defined_names_fixture(path: pathlib.Path) -> None:
    workbook = Workbook()
    apply_deterministic_metadata(workbook, "defined names fixture")
    sheet = workbook.active
    sheet.title = "Sheet1"
    sheet.append(["Amount"])
    for amount in [100, 200, 300, 400]:
        sheet.append([amount])
    sheet["C1"] = "Total"
    sheet["C2"] = "=SUM(SalesRange)"
    sheet["C3"] = "=C2*TaxRate"
    workbook.defined_names.add(
        DefinedName(name="SalesRange", attr_text="Sheet1!$A$2:$A$5")
    )
    workbook.defined_names.add(DefinedName(name="TaxRate", attr_text="0.07"))
    workbook.save(path)


def generate_descriptors(output_dir: pathlib.Path) -> list[FixtureDescriptor]:
    output_dir.mkdir(parents=True, exist_ok=True)
    created: list[FixtureDescriptor] = []

    minimal_sheet = (
        '<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
        "<sheetData><row r=\"1\"><c r=\"A1\"><v>1</v></c></row></sheetData>"
        "</worksheet>"
    )
    minimal_path = output_dir / "minimal.xlsx"
    write_fixture(minimal_path, minimal_sheet)
    created.append(
        FixtureDescriptor(
            name="minimal",
            path=minimal_path,
            features=("values",),
            generator="manual_xml",
        )
    )

    formula_sheet = (
        '<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
        "<sheetData>"
        "<row r=\"1\"><c r=\"A1\"><v>10</v></c></row>"
        "<row r=\"2\"><c r=\"A2\"><v>2</v></c></row>"
        "<row r=\"3\"><c r=\"A3\"><f>A1+A2</f><v>12</v></c></row>"
        "</sheetData>"
        "</worksheet>"
    )
    formula_path = output_dir / "formula.xlsx"
    write_fixture(formula_path, formula_sheet)
    created.append(
        FixtureDescriptor(
            name="formula",
            path=formula_path,
            features=("formula",),
            generator="manual_xml",
        )
    )

    dangling_sheet = (
        '<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
        "<sheetData><row r=\"1\"><c r=\"A1\"><v>7</v></c></row></sheetData>"
        "</worksheet>"
    )
    dangling_path = output_dir / "dangling-edge.xlsx"
    write_fixture(dangling_path, dangling_sheet, extra_dangling=True)
    created.append(
        FixtureDescriptor(
            name="dangling-edge",
            path=dangling_path,
            features=("dangling_relationship",),
            generator="manual_xml",
        )
    )

    unknown_part_sheet = (
        '<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
        "<sheetData><row r=\"1\"><c r=\"A1\"><v>99</v></c></row></sheetData>"
        "</worksheet>"
    )
    unknown_path = output_dir / "unknown-part.xlsx"
    write_fixture(
        unknown_path,
        unknown_part_sheet,
        extra_parts={"xCustom/blob.bin": b"opaque"},
    )
    created.append(
        FixtureDescriptor(
            name="unknown-part",
            path=unknown_path,
            features=("unknown_part",),
            generator="manual_xml",
        )
    )

    styles_path = output_dir / "styles.xlsx"
    write_styles_fixture(styles_path)
    created.append(
        FixtureDescriptor(
            name="styles",
            path=styles_path,
            features=("styles", "formatting"),
            generator="openpyxl",
        )
    )

    comments_path = output_dir / "comments.xlsx"
    write_comments_fixture(comments_path)
    created.append(
        FixtureDescriptor(
            name="comments",
            path=comments_path,
            features=("comments",),
            generator="openpyxl",
        )
    )

    chart_path = output_dir / "chart.xlsx"
    write_chart_fixture(chart_path)
    created.append(
        FixtureDescriptor(
            name="chart",
            path=chart_path,
            features=("charts",),
            generator="openpyxl",
        )
    )

    defined_names_path = output_dir / "defined-names.xlsx"
    write_defined_names_fixture(defined_names_path)
    created.append(
        FixtureDescriptor(
            name="defined-names",
            path=defined_names_path,
            features=("defined_names",),
            generator="openpyxl",
        )
    )

    return created


def generate(output_dir: pathlib.Path) -> list[pathlib.Path]:
    return [descriptor.path for descriptor in generate_descriptors(output_dir)]


def write_manifest(
    output_dir: pathlib.Path, manifest_path: pathlib.Path, descriptors: list[FixtureDescriptor]
) -> None:
    manifest = {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "output_dir": str(output_dir.resolve()),
        "fixture_count": len(descriptors),
        "fixtures": [
            {
                "name": descriptor.name,
                "path": str(descriptor.path.resolve()),
                "relative_path": descriptor.path.name,
                "features": list(descriptor.features),
                "generator": descriptor.generator,
            }
            for descriptor in descriptors
        ],
    }
    manifest_path.parent.mkdir(parents=True, exist_ok=True)
    manifest_path.write_text(json.dumps(manifest, indent=2), encoding="utf-8")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Generate deterministic XLSX fixtures for corpus validation."
    )
    parser.add_argument(
        "--output-dir",
        default=".ci/corpus-fixtures",
        help="Directory to write generated .xlsx fixtures.",
    )
    parser.add_argument(
        "--manifest",
        default=None,
        help=(
            "Optional manifest path. Defaults to <output-dir>/manifest.json. "
            "Use --no-manifest to disable."
        ),
    )
    parser.add_argument(
        "--no-manifest",
        action="store_true",
        help="Disable writing fixture metadata manifest.",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    output_dir = pathlib.Path(args.output_dir)
    descriptors = generate_descriptors(output_dir)
    generated = [descriptor.path for descriptor in descriptors]

    manifest_path: pathlib.Path | None = None
    if not args.no_manifest:
        manifest_path = (
            pathlib.Path(args.manifest)
            if args.manifest
            else output_dir / "manifest.json"
        )
        write_manifest(output_dir, manifest_path, descriptors)

    print(f"Generated {len(generated)} fixture(s) in {output_dir}")
    for descriptor in descriptors:
        feature_list = ",".join(descriptor.features)
        print(
            f" - {descriptor.path} [{descriptor.generator}] "
            f"features={feature_list}"
        )
    if manifest_path is not None:
        print(f"Manifest: {manifest_path}")


if __name__ == "__main__":
    main()
